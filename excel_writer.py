from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

RESULT_PATH = Path("wyniki_analizy.xlsx")
HEADERS = [
    "hotel",
    "check_in",
    "check_out",
    "status",
    "rack_price",
    "package_price",
    "page_title",
    "notes",
]


def init_workbook() -> None:
    if RESULT_PATH.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Analiza"
    ws.append(HEADERS)
    wb.save(RESULT_PATH)


def append_result(row_data: dict) -> None:
    init_workbook()
    wb = load_workbook(RESULT_PATH)
    ws = wb["Analiza"]
    ws.append([row_data.get(header) for header in HEADERS])
    wb.save(RESULT_PATH)
